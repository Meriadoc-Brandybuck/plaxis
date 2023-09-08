import os
from plxscripting.easy import*
import openpyxl
from collections.abc import MutableMapping
import plxscripting as plc

excel = 2
path = 0
names = os.listdir('G:\\plaxis2d\\plaxis2d')
wb = openpyxl.load_workbook(
        'C:\\Users\\DELL\\PycharmProjects\\pythonProject\\临隧基坑开挖\\新建 Microsoft Excel 工作表.xlsx')
ws = wb['Sheet1']
for i in names:
    s_o, g_o = new_server('localhost', 10000, password='e6STEwUwe!Mw#saB')
    kid = i.split('.')[0]
    kid1 = i.split('.')[-1]
    if kid1 == 'p2dx':
        path = "G:\plaxis2d\plaxis2d\{}". format(i)
        path = 'r'+"'{}'".format(path) + ','
        print(path)

a = [r'G:\plaxis2d\plaxis2d\12-12.p2dx',
r'G:\plaxis2d\plaxis2d\12-15.p2dx',
r'G:\plaxis2d\plaxis2d\12-18.p2dx',
r'G:\plaxis2d\plaxis2d\12-21.p2dx',
r'G:\plaxis2d\plaxis2d\12-27.p2dx',
r'G:\plaxis2d\plaxis2d\12-36.p2dx',
r'G:\plaxis2d\plaxis2d\12-45.p2dx',
r'G:\plaxis2d\plaxis2d\12-9.p2dx',
r'G:\plaxis2d\plaxis2d\15-12.p2dx',
r'G:\plaxis2d\plaxis2d\15-15.p2dx',
r'G:\plaxis2d\plaxis2d\15-18.p2dx',
r'G:\plaxis2d\plaxis2d\15-21.p2dx',
r'G:\plaxis2d\plaxis2d\15-27.p2dx',
r'G:\plaxis2d\plaxis2d\15-36.p2dx',
r'G:\plaxis2d\plaxis2d\15-45.p2dx',
r'G:\plaxis2d\plaxis2d\15-9.p2dx',
r'G:\plaxis2d\plaxis2d\21-12.p2dx',
r'G:\plaxis2d\plaxis2d\21-15.p2dx',
r'G:\plaxis2d\plaxis2d\21-18.p2dx',
r'G:\plaxis2d\plaxis2d\21-21.p2dx',
r'G:\plaxis2d\plaxis2d\21-27.p2dx',
r'G:\plaxis2d\plaxis2d\21-36.p2dx',
r'G:\plaxis2d\plaxis2d\21-45.p2dx',
r'G:\plaxis2d\plaxis2d\21-9.p2dx',
r'G:\plaxis2d\plaxis2d\27-12.p2dx',
r'G:\plaxis2d\plaxis2d\27-15.p2dx',
r'G:\plaxis2d\plaxis2d\27-18.p2dx',
r'G:\plaxis2d\plaxis2d\27-21.p2dx',
r'G:\plaxis2d\plaxis2d\27-27.p2dx',
r'G:\plaxis2d\plaxis2d\27-36.p2dx',
r'G:\plaxis2d\plaxis2d\27-45.p2dx',
r'G:\plaxis2d\plaxis2d\27-9.p2dx',
r'G:\plaxis2d\plaxis2d\33-12.p2dx',
r'G:\plaxis2d\plaxis2d\33-15.p2dx',
r'G:\plaxis2d\plaxis2d\33-18.p2dx',
r'G:\plaxis2d\plaxis2d\33-21.p2dx',
r'G:\plaxis2d\plaxis2d\33-27.p2dx',
r'G:\plaxis2d\plaxis2d\33-36.p2dx',
r'G:\plaxis2d\plaxis2d\33-45.p2dx',
r'G:\plaxis2d\plaxis2d\33-9.p2dx',
r'G:\plaxis2d\plaxis2d\39-12.p2dx',
r'G:\plaxis2d\plaxis2d\39-15.p2dx',
r'G:\plaxis2d\plaxis2d\39-18.p2dx',
r'G:\plaxis2d\plaxis2d\39-21.p2dx',
r'G:\plaxis2d\plaxis2d\39-27.p2dx',
r'G:\plaxis2d\plaxis2d\39-36.p2dx',
r'G:\plaxis2d\plaxis2d\39-45.p2dx',
r'G:\plaxis2d\plaxis2d\39-9.p2dx',
r'G:\plaxis2d\plaxis2d\6-12.p2dx',
r'G:\plaxis2d\plaxis2d\6-15.p2dx',
r'G:\plaxis2d\plaxis2d\6-18.p2dx',
r'G:\plaxis2d\plaxis2d\6-21.p2dx',
r'G:\plaxis2d\plaxis2d\6-27.p2dx',
r'G:\plaxis2d\plaxis2d\6-36.p2dx',
r'G:\plaxis2d\plaxis2d\6-45.p2dx',
r'G:\plaxis2d\plaxis2d\6-9.p2dx',
r'G:\plaxis2d\plaxis2d\9-12.p2dx',
r'G:\plaxis2d\plaxis2d\9-15.p2dx',
r'G:\plaxis2d\plaxis2d\9-18.p2dx',
r'G:\plaxis2d\plaxis2d\9-21.p2dx',
r'G:\plaxis2d\plaxis2d\9-27.p2dx',
r'G:\plaxis2d\plaxis2d\9-36.p2dx',
r'G:\plaxis2d\plaxis2d\9-45.p2dx',
r'G:\plaxis2d\plaxis2d\9-9.p2dx']


'''
以下for循环用于得出某一阶段的计算结果，排桩和隧道都是板结构，通过内部序号分辨，土工格栅通过位置和名称分辨
如果基坑宽度有变化，需要修改土工格栅处的x_left（基坑右侧x值）
'''

for i in a:
    s_o, g_o = new_server('localhost', 10000, password='e6STEwUwe!Mw#saB')
    s_o.open(i)
    plateX = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.X, 'node')
    plateY = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.Y, 'node')
    plateUx = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.Ux, 'node')
    plateUy = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.Uy, 'node')
    plateMAT = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.MaterialID, 'node')
    minUx = 0.0
    maxUx = 0.0
    xAtMaxUx = 0.0
    yAtMaxUx = 0.0
    xAtMinUx = 0.0
    yAtMinUx = 0.0
    maxUy = 0.0
    minUy = 0.0
    xAtMaxUy = 0.0
    yAtMaxUy = 0.0
    xAtMinUy = 0.0
    yAtMinUy = 0.0
    for x, y, ux, uy, mat in zip(plateX, plateY, plateUx, plateUy, plateMAT):
        if mat == 1:
            if ux < minUx:
                minUx = ux
                xAtMinUx = x
                yAtMinUx = y
            elif ux > maxUx:
                maxUx = ux
                xAtMaxUx = x
                xAtMaxUx = y
            if uy < minUy:
                minUy = uy
                xAtMinUy = x
                yAtMinUy = y
            elif uy > maxUy:
                maxUy = uy
                xAtMaxUy = x
                xAtMaxUy = y
            if maxUx > -minUx:
                print("隧道最大水平位移：{:.6f}m".format(maxUx))
                ws.cell(row=excel, column=3).value = format(maxUx)
            else:
                print("隧道最大水平位移：{:.6f}m".format(minUx))
                ws.cell(row=excel, column=3).value = format(minUx)
            if maxUy > -minUy:
                print("隧道最大竖向位移：{:.6f}m".format(maxUy))
                ws.cell(row=excel, column=4).value = format(maxUy)
            else:
                print("隧道最大竖向位移：{:.6f}m".format(minUy))
                ws.cell(row=excel, column=4).value = format(minUy)
    minUx = 0
    maxUx = 0.0
    xAtMaxUx = 0.0
    yAtMaxUx = 0.0
    xAtMinUx = 0.0
    yAtMinUx = 0.0
    plateX = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.X, 'node')
    plateY = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.Y, 'node')
    plateUx = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.Plate.Ux, 'node')
    for x, y, ux, uy, mat in zip(plateX, plateY, plateUx, plateUy, plateMAT):
        if mat == 2:
            if ux < minUx:
                minUx = ux
                xAtMinUx = x
                yAtMinUx = y
            elif ux > maxUx:
                maxUx = ux
                xAtMaxUx = x
                xAtMaxUx = y
    if maxUx > -minUx:
        print("排桩最大水平位移：{:.6f}m".format(maxUx))
        ws.cell(row=excel, column=5).value = format(maxUx)
    else:
        print("排桩最大水平位移：{:.6f}m".format(minUx))
        ws.cell(row=excel, column=5).value = format(minUx)
    x_left = 15.0
    x_right = 100.0
    y_bottom = 0.0
    maxUy = 0.0
    minUy = 0.0
    xAtMaxUy = 0.0
    yAtMaxUy = 0.0
    geogridX = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.GeoGrid.X, 'node')
    geogridY = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.GeoGrid.Y, 'node')
    geogridUy = g_o.getresults(g_o.Phase_5, g_o.ResultTypes.GeoGrid.Uy, 'node')
    for x, y, uy in zip(geogridX, geogridY, geogridUy):
        if x_left < x < x_right:
            if abs(y - y_bottom) < 1E-5:
                if uy > maxUy:
                    maxUy = uy
                    xAtMaxUy = x
                    yAtMaxUy = y
                elif uy < minUy:
                    minUy = uy
                    xAtMaxUy = x
                    xAtMaxUy = y
    if maxUy > -minUy:
        print("土工格栅最大竖向位移：{:.6f}m".format(maxUy))
        ws.cell(row=excel, column=6).value = format(maxUy)
    else:
        print("土工格栅最大竖向位移：{:.6f}m".format(minUy))
        ws.cell(row=excel, column=6).value = format(minUy)
    s_o.close()
    ws.cell(row=excel, column=1).value = i
    excel += 1

excel = 2
for i in a:
    s_o, g_o = new_server('localhost', 10000, password='e6STEwUwe!Mw#saB')
    s_o.open(i)
    plateX = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.X, 'node')
    plateY = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.Y, 'node')
    plateUx = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.Ux, 'node')
    plateUy = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.Uy, 'node')
    plateMAT = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.MaterialID, 'node')
    minUx = 0.0
    maxUx = 0.0
    xAtMaxUx = 0.0
    yAtMaxUx = 0.0
    xAtMinUx = 0.0
    yAtMinUx = 0.0
    maxUy = 0.0
    minUy = 0.0
    xAtMaxUy = 0.0
    yAtMaxUy = 0.0
    xAtMinUy = 0.0
    yAtMinUy = 0.0
    for x, y, ux, uy, mat in zip(plateX, plateY, plateUx, plateUy, plateMAT):
        if mat == 1:
            if ux < minUx:
                minUx = ux
                xAtMinUx = x
                yAtMinUx = y
            elif ux > maxUx:
                maxUx = ux
                xAtMaxUx = x
                xAtMaxUx = y
            if uy < minUy:
                minUy = uy
                xAtMinUy = x
                yAtMinUy = y
            elif uy > maxUy:
                maxUy = uy
                xAtMaxUy = x
                xAtMaxUy = y
            if maxUx > -minUx:
                print("隧道最大水平位移：{:.6f}m".format(maxUx))
                ws.cell(row=excel, column=8).value = format(maxUx)
            else:
                print("隧道最大水平位移：{:.6f}m".format(minUx))
                ws.cell(row=excel, column=8).value = format(minUx)
            if maxUy > -minUy:
                print("隧道最大竖向位移：{:.6f}m".format(maxUy))
                ws.cell(row=excel, column=9).value = format(maxUy)
            else:
                print("隧道最大竖向位移：{:.6f}m".format(minUy))
                ws.cell(row=excel, column=9).value = format(minUy)
    minUx = 0
    maxUx = 0.0
    xAtMaxUx = 0.0
    yAtMaxUx = 0.0
    xAtMinUx = 0.0
    yAtMinUx = 0.0
    plateX = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.X, 'node')
    plateY = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.Y, 'node')
    plateUx = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.Plate.Ux, 'node')
    for x, y, ux, uy, mat in zip(plateX, plateY, plateUx, plateUy, plateMAT):
        if mat == 2:
            if ux < minUx:
                minUx = ux
                xAtMinUx = x
                yAtMinUx = y
            elif ux > maxUx:
                maxUx = ux
                xAtMaxUx = x
                xAtMaxUx = y
    if maxUx > -minUx:
        print("排桩最大水平位移：{:.6f}m".format(maxUx))
        ws.cell(row=excel, column=10).value = format(maxUx)
    else:
        print("排桩最大水平位移：{:.6f}m".format(minUx))
        ws.cell(row=excel, column=10).value = format(minUx)
    x_left = 15.0
    x_right = 100.0
    y_bottom = 0.0
    maxUy = 0.0
    minUy = 0.0
    xAtMaxUy = 0.0
    yAtMaxUy = 0.0
    geogridX = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.GeoGrid.X, 'node')
    geogridY = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.GeoGrid.Y, 'node')
    geogridUy = g_o.getresults(g_o.Phase_6, g_o.ResultTypes.GeoGrid.Uy, 'node')
    for x, y, uy in zip(geogridX, geogridY, geogridUy):
        if x_left < x < x_right:
            if abs(y - y_bottom) < 1E-5:
                if uy > maxUy:
                    maxUy = uy
                    xAtMaxUy = x
                    yAtMaxUy = y
                elif uy < minUy:
                    minUy = uy
                    xAtMaxUy = x
                    xAtMaxUy = y
    if maxUy > -minUy:
        print("土工格栅最大竖向位移：{:.6f}m".format(maxUy))
        ws.cell(row=excel, column=11).value = format(maxUy)
    else:
        print("土工格栅最大竖向位移：{:.6f}m".format(minUy))
        ws.cell(row=excel, column=11).value = format(minUy)
    s_o.close()
    ws.cell(row=excel, column=1).value = i
    excel += 1

ws.cell(row=1, column=1).value = '名称'
ws.cell(row=1, column=3).value = '隧道最大水平位移_Phase_5'
ws.cell(row=1, column=4).value = '隧道最大竖向位移_Phase_5'
ws.cell(row=1, column=5).value = '排桩最大水平位移_Phase_5'
ws.cell(row=1, column=6).value = '土工格栅最大竖向位移_Phase_5'
ws.cell(row=1, column=8).value = '隧道最大水平位移_Phase_6'
ws.cell(row=1, column=9).value = '隧道最大竖向位移_Phase_6'
ws.cell(row=1, column=10).value = '排桩最大水平位移_Phase_6'
ws.cell(row=1, column=11).value = '土工格栅最大竖向位移_Phase_6'
wb.save('C:\\Users\\DELL\\PycharmProjects\\pythonProject\\临隧基坑开挖\\新建 Microsoft Excel 工作表.xlsx')